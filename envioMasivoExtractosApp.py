# -*- coding: utf-8 -*-
"""
Created on Mon Nov 11 13:51:16 2024

@author: angperilla
"""


"""
    Script para el proceso de generar y enviar extractos masivos de todas las 
    IPS de acuerdo al mes de actualización.
    
    Insumos: 
        1. HR filtrada por F.Aviso del mes a Trabajar (xlsx) - HR Analítica -
        2. CONTACTO PRESTADORES (xlsx) - Base de Correos -
        3. BG (jpg) - Imagen Fondo -  
          
    Salidas:
        1. Destinatarios (xlsx) - Relación de Extractos vs Correos Destinatarios -
        2. Resultado Envío Correos (xlsx) - Status Envío correos -
        
"""

#############################################################################
############################ Librerías Proyecto #############################
#############################################################################

# --------------------- Librerias Extractos Masivos -------------------------

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import datetime, timedelta


import os
import glob
import re

# --------------------- Librerias Envio Correos Masivos ---------------------

import win32com.client

# ------------------------------- Interfaz ----------------------------------
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import time

#----------------------------------------------------------------------------

#############################################################################



# Función para contar archivos Excel en el directorio EXCEL
def contar_archivos_excel():
    ruta_excel = os.path.join(os.getcwd(), 'EXCEL')
    if not os.path.exists(ruta_excel):
        return 0  # Si no existe la carpeta EXCEL
    archivos_excel = [f for f in os.listdir(ruta_excel) if f.endswith('.xlsx')]
    return len(archivos_excel)



# Función para contar registros en el archivo Destinatarios.xlsx
def contar_destinatarios():
    ruta_destinatarios = os.path.join(os.getcwd(), 'data', 'Destinatarios.xlsx')
    if not os.path.exists(ruta_destinatarios):
        return 0  # Si no existe el archivo de destinatarios
    df_destinatarios = pd.read_excel(ruta_destinatarios)
    # Filtrar destinatarios con correos válidos
    df_destinatarios = df_destinatarios[df_destinatarios['CorreoPrincipal'].notna()]
    
    return len(df_destinatarios)



def generar_extractos():
    try:
    
        # inicio = time.time()
        # Obtener la ruta actual
        ruta_actual = os.getcwd()
    
    
        def listar_archivos_xlsx_en_data(ruta_actual):
            """
                Lista todos los archivos .xlsx en la carpeta 'data' dentro de la ruta actual.
                
                Returns:
                list: Una lista de nombres de archivos .xlsx encontrados en la carpeta 'data'.
                
            """
            
            # Concatenar con la carpeta 'data'
            ruta_data = os.path.join(ruta_actual, 'data')
            
            # Verificar si la carpeta 'data' existe
            if not os.path.exists(ruta_data):
                print("La carpeta 'data' no existe.")
                return []
            
            # Buscar todos los archivos .xlsx en el directorio 'data'
            archivos_xlsx = glob.glob(os.path.join(ruta_data, "*.xlsx"))
           
            # Devolver la lista de archivos .xlsx encontrados
            return archivos_xlsx
    
    
        def extraer_mes_de_archivo(archivo):
            """
            Extrae los dos dígitos antes del guion bajo (_) del nombre del archivo.
    
            Args:
                archivo (str): El nombre del archivo.
    
            Returns:
                str: Los dos dígitos extraídos del nombre del archivo.
            """
            match = re.search(r'(\d{2})_', archivo)
            if match:
                return match.group(1)
            else:
                return None
    
    
        # Listar archivos .xlsx en la carpeta 'data'
        archivos = listar_archivos_xlsx_en_data(ruta_actual)
    
    
        for archivo in archivos:
            if 'Hoja Ruta' in archivo:
                mes = extraer_mes_de_archivo(os.path.basename(archivo))
                ruta = archivo
            
                if mes:
                    print(f"El mes extraído del archivo '{archivo}' es: {mes}")
                else:
                    print(f"El archivo no está nombrado correctamente: '{archivo}'.")
            else:
                print("Debe estar la Hoja de Ruta del mes en la carpeta 'data'.")
    
    
        #__________________________ Extraer Información de Fechas____________________
    
    
        # Diccionario para traducir los nombres de los meses
        meses = {
            "January": "Enero", "February": "Febrero", "March": "Marzo", "April": "Abril",
            "May": "Mayo", "June": "Junio", "July": "Julio", "August": "Agosto",
            "September": "Septiembre", "October": "Octubre", "November": "Noviembre", "December": "Diciembre"
        }
    
    
        # Año
        año = 2024
    
        # Crear la variable fecha_corte
        fecha_corte = datetime(año, int(mes), 1).strftime("%B de %Y")
        fecha_corte = fecha_corte.replace(fecha_corte.split()[0], meses[fecha_corte.split()[0]])
    
    
        # Crear la variable ultima_fecha_mes
        ultimo_dia_mes = (datetime(año, int(mes) + 1, 1) - timedelta(days=1)).strftime("%d de %B de %Y")
        ultimo_dia_mes = ultimo_dia_mes.replace(ultimo_dia_mes.split()[2], meses[ultimo_dia_mes.split()[2]])
    
    
        # print("fecha_corte =", fecha_corte)
        # print("ultima_fecha_mes =", ultimo_dia_mes)
    
    
        #____________________________________________________________________________
    
    
    
        # Lectura HR
        data_hr = pd.read_excel(ruta)
    
    
        # Valores estado factura
        # estado_hr = list(data_hr['ESTADO ACTUAL FACTURA'].unique())
        # print(estado_hr)
    
    
        # # Obtener los valores únicos de la columna 
        # valores_amparos = data_hr['AMPARO'].unique()
    
    
        # # Valores que no cruzan por Ampara los Nan
        # amparos_nan = data_hr[data_hr['AMPARO'].isna()]
    
    
        # Filtrar el DataFrame por los valores específicos en la columna 'AMPARO'
        amparos = ['GASTOS MEDICOS, QUIRURGICOS, FARMACEUTICOS Y HOSPI', 
                           'GASTOS DE TRANSPORTE Y MOVILIZACION DE VICTIMAS']
    
    
        # # Filtrar HR por amparo 
        # data_hr = data_hr[data_hr['AMPARO'].isin(amparos)]
    
    
        # Filtar HR por estados validos
        data_hr = data_hr[~data_hr['ESTADO ACTUAL FACTURA'].str.contains("ERROR", na=False)]
        data_hr = data_hr[data_hr['AMPARO'].isin(amparos) | data_hr['AMPARO'].isna()]
    
    
        # # Información muestra
        # id1 = data_hr[data_hr['ID RECLAMANTE'] == 900431550]
        # # id1.to_excel('resultado.xlsx', index=False)
        # id2 = data_hr[data_hr['ID RECLAMANTE'] == 891401643]
    
    
        # # @todo
        # # Utilizar nits de prubeba
        # # Destinatarios
        nits = list(data_hr['ID RECLAMANTE'].unique())
        # print(len(nits)) # 1405 Nits para Mayo
    
        # 
        # nits = [900006037]
    
        # Columnas HR
        columnas_hr = ['NUMERO FACTURA',
                       'F.AVISO',
                       'SINIESTRO',
                       'FACTURA IQ',
                       'ID RECLAMANTE', # En el resultado final no se tiene en cuenta
                       'RECLAMANTE',
                       'VLR RADICACION',
                       'VLR APROBADO',
                       'VLR GLOSADO',
                       'ESTADO ACTUAL FACTURA',
                       'F.GIRO']
    
    
        # Columnas Extracto
        columnas_extracto = [
            "NUMERO_FACTURA",
            "FECHA_AVISO",
            "SINIESTRO",
            "NUMERO_DE_RADICADO",
            'ID RECLAMANTE', # En el resultado final no se tiene en cuenta
            'RECLAMANTE',
            "VALOR_RECLAMADO", 
            "VALOR_PAGADO",
            "VALOR_OBJECION",
            "ESTADO",
            "ESTADO_DE_PAGO",
            "FECHA_GIRO"]
    
    
        # Filtrar las columnas que no se necesitan
        columnas_extracto_filtradas = [col for col in columnas_extracto if col != 'ESTADO_DE_PAGO']
    
    
        # Crear un diccionario para el mapeo de nombres
        columnas_renombradas = dict(zip(columnas_hr, columnas_extracto_filtradas))
    
    
        # HR Renombrada
        data_hr_renombrada = data_hr[columnas_hr].rename(columns=columnas_renombradas)
        # print(data_hr_renombrada.info())
    
    
        def determinar_estado(row):
            """
                De acuerdo a valores de las columnas: VALOR_RECLAMADO, VALOR_PAGADO y
                VALOR_OBJECION determina el ESTADO de la factura.
                
                Output:
                - 'Pago Con Objecion Parcial'
                - 'Pago Total'
                - 'Objecion Parcial'
                - 'Objecion Total/Devolucion'
                
            """
            estados_liquidado_con_pago = ['LIQUIDADO CON PAGO']
            estados_liquidado_sin_pago = ['LIQUIDADO SIN PAGO']
            estados_comunicacion = [
                'COMUNICACIÓN ENVIADA POR OBJECIÓN',
                'COMUNICACIÓN ENVIADA POR DEVOLUCIÓN',
                'COMUNICACIÓN ENVIADA POR DEVOLUCIÓN-LIQUIDADO SIN',
                'COMUNICACIÓN ENVIADA POR OBJECIÓN-LIQUIDADO SIN PA'
            ]
    
            if row['VALOR_RECLAMADO'] > 0:
                if row['VALOR_PAGADO'] > 0:
                    if row['VALOR_OBJECION'] > 0 and row['ESTADO'] in estados_liquidado_con_pago:
                        return 'Pago Con Objecion Parcial'
                    elif row['VALOR_OBJECION'] == 0 and row['ESTADO'] in estados_liquidado_con_pago:
                        return 'Pago Total'
                elif row['VALOR_PAGADO'] == 0:
                    if row['VALOR_OBJECION'] > 0:
                        if row['ESTADO'] in estados_liquidado_sin_pago:
                            return 'Objecion Parcial'
                        elif row['ESTADO'] in estados_comunicacion:
                            return 'Objecion Total/Devolucion'
            elif row['VALOR_RECLAMADO'] == 0:
                if row['VALOR_PAGADO'] == 0 and row['VALOR_OBJECION'] == 0:
                    if row['ESTADO'] in estados_comunicacion: 
                        return 'Objecion Total/Devolucion'
                        
            return 'Objecion Parcial'
    
    
        # Crear un diccionario vacío para almacenar los datos
        data_dict = {'ID RECLAMANTE': [], 'RECLAMANTE': [], 'NOMBRE EXTRACTO IPS': [], 'FECHA CORTE': [], 'FECHA ULTIMO DIA MES':[]}
    
    
        # Ciclo para recorrer cada Nit y transformar información
        for nit in nits:
            # Filtrar por nit
            df_filtered = data_hr_renombrada[data_hr_renombrada['ID RECLAMANTE'] == nit].copy()
            
            # Obtener Nombre Entidad
            entidad = df_filtered['RECLAMANTE'].iloc[0]
            
            # Corregir la columna 'VALOR_OBJECION' solo si 'VALOR_PAGADO' es vacía
            df_filtered.loc[df_filtered['VALOR_PAGADO'].isna(), 'VALOR_OBJECION'] = df_filtered['VALOR_RECLAMADO'] - df_filtered['VALOR_PAGADO'].fillna(0)
            
            # Asegurarse de que 'VALOR_PAGADO' quede con valor de 0 si es vacío
            df_filtered.loc[:,'VALOR_PAGADO'] = df_filtered['VALOR_PAGADO'].fillna(0)
            
            # Determinar el estado
            df_filtered.loc[:,'ESTADO_FINAL'] = df_filtered.apply(determinar_estado, axis=1)
            
            # Crear la columna ESTADO_DE_PAGO
            df_filtered.loc[:,'ESTADO_DE_PAGO'] = df_filtered.apply(
                lambda row: 'Giro Realizado' if pd.notna(row['FECHA_GIRO']) else ('Giro Por Realizar' if row['ESTADO'] == 'LIQUIDADO CON PAGO' else ''),
                axis=1
            )
            
            # print(df_filtered)
                
            # Actualizar informacion ESTADO
            df_filtered.loc[:,'ESTADO'] = df_filtered['ESTADO_FINAL']
            
            # Seleccion de columnas
            df_filtered = df_filtered[columnas_extracto]
            
            
            # Convertir la columna 'FECHA_GIRO' a datetime
            df_filtered['FECHA_GIRO'] = pd.to_datetime(df_filtered['FECHA_GIRO'], errors='coerce')
            df_filtered['FECHA_AVISO'] = pd.to_datetime(df_filtered['FECHA_AVISO'], errors='coerce')
    
            # Formatear Fechas
            # Verificar si la conversión fue exitosa
            if not df_filtered['FECHA_GIRO'].isnull().any():
                df_filtered.loc[:, 'FECHA_GIRO'] = df_filtered['FECHA_GIRO'].dt.strftime('%Y-%m-%d')
            
            
            if not df_filtered['FECHA_AVISO'].isnull().any():
                df_filtered.loc[:, 'FECHA_AVISO'] = df_filtered['FECHA_AVISO'].dt.strftime('%Y-%m-%d')
            
            # Eliminar Columnas del resultado
            df_filtered = df_filtered.drop(columns=['ID RECLAMANTE'])
            df_filtered = df_filtered.drop(columns=['RECLAMANTE'])
            
            # Ordenar por 'FECHA_AVISO' y 'NUMERO_FACTURA'
            df_filtered = df_filtered.sort_values(by=['FECHA_AVISO', 'NUMERO_FACTURA'], ascending=[False, True])
            
            # df_filtered.info()
              
            
            df_transformed = df_filtered.copy()
    
            # Intentar crear un Workbook
            try:
                wb = Workbook()
                # print("Workbook creado exitosamente.")
                # print(f"Tipo de wb: {type(wb)}")
            except Exception as e:
                print(f"Error al crear el Workbook: {e}")
                
            # # Verificar si wb es None
            # if wb is None:
            #     print("El objeto Workbook es None.")
            # else:
            #     print("El objeto Workbook es válido.")
            
            # Eliminar la hoja predeterminada
            wb.remove(wb.active)
    
            
            # Crear una tabla dinámica (pivot table) usando conteo
            pivot_table = df_filtered.pivot_table(index=['ESTADO', 'ESTADO_DE_PAGO'], values='NUMERO_DE_RADICADO', aggfunc='count', margins=True, margins_name='TOTAL').reset_index()
            
            # Renombrar la columna
            pivot_table.rename(columns={'NUMERO_DE_RADICADO': 'CANTIDAD_RECLAMACIONES'}, inplace=True)
            pivot_table = pivot_table.sort_values(by=['ESTADO_DE_PAGO'], ascending=[False])
            
            # # Calcular el total
            # total = pivot_table['CANTIDAD_RECLAMACIONES'].sum()
            
            # # Crear una fila de vacio
            # fila_nula = pd.DataFrame({'ESTADO': [''], 'ESTADO_DE_PAGO': [''], 'CANTIDAD_RECLAMACIONES': ['']})
            
            # # Crear una fila de total
            # total_row = pd.DataFrame({'ESTADO': ['TOTAL'], 'ESTADO_DE_PAGO': [''], 'CANTIDAD_RECLAMACIONES': [total]})
            
            # # Agregar la fila de total a la tabla
            # pivot_table = pd.concat([pivot_table, fila_nula, total_row], ignore_index=True)
    
            # print(pivot_table)
            
            # # Aplicar estilo a la tabla
            # styled_table = pivot_table.style.set_table_styles(
            #     [{'selector': 'th', 'props': [('background-color', '#f7f7f9'), ('color', '#333'), ('border', '1px solid #ddd')]},
            #      {'selector': 'td', 'props': [('border', '1px solid #ddd')]}]
            # ).set_properties(**{'text-align': 'center'})
            
            # # Mostrar la tabla con estilo
            # styled_table
            
            # grouped_df = df_filtered.groupby(['ESTADO', 'ESTADO_DE_PAGO']).size().reset_index(name='NUMERO_DE_RADICADO')
            
            
            ##########################################################################
            #________________________ Exportar Excel Hoja Resumen ____________________
            ##########################################################################
            
            
            # Crear una nueva hoja para la tabla dinámica
            ws1 = wb.create_sheet(title="Resumen", index=0)
            
            
            #___________________ Texto Combinado Hoja Resumen________________________
            # Combinar las celdas C7, D7 y E7 en la hoja 'Resumen'
            ws1.merge_cells('C7:E7')
            
            # Establecer el texto en la celda combinada
            cell = ws1['C7']
            cell.value = 'EXTRACTO DE RECLAMACIONES - RESUMEN'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='258BFB', end_color='258BFB', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            #________________________________________________________________________
            
            
            # Escribir la tabla dinámica en la nueva hoja a partir de la celda C9
            for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), 9):
                for c_idx, value in enumerate(row, 3):
                    cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
            # Aplicar estilos a los encabezados
            header_fill = PatternFill(start_color='258BFB', end_color='258BFB', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Aplicar estilos a los encabezados en la nueva hoja desde la celda C9
            for cell in ws1[9][2:]:  # ws2[9] es la fila 9, [2:] omite las primeras dos celdas (A9 y B9)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Quitar cuadrículas
            ws1.sheet_view.showGridLines = False
            
            # Autoajustar las columnas
            for col in ws1.iter_cols(min_col=3):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws1.column_dimensions[column].width = adjusted_width
                
                
                
            #########################################################################
            #_____________________Exportar Excel Hoja Detalle________________________
            #########################################################################
            
            
            # Crear una nueva hoja para el detalle
            ws2 = wb.create_sheet(title="Detalle", index=1)
            
            
            #___________________ Texto Combinado Hoja Detalle________________________
            # Combinar las celdas E4, F4, G4 y H4 en la hoja 'Resumen'
            ws2.merge_cells('E4:H4')
            
            # Establecer el texto en la celda combinada
            cell = ws2['E4']
            cell.value = 'EXTRACTO DE FACTURAS - DETALLE'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='258BFB', end_color='258BFB', fill_type='solid')
            cell.font = Font(color='FFFFFF', bold=True)
            #________________________________________________________________________
            
    
            # Escribir el DataFrame en el archivo Excel a partir de la celda B7
            for r_idx, row in enumerate(dataframe_to_rows(df_transformed, index=False, header=True), 7):
                for c_idx, value in enumerate(row, 2):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    
            for cell in ws2[7][1:]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
             
            # Quitar cuadrículas
            ws2.sheet_view.showGridLines = False 
            
            # Autoajustar las columnas
            for col in ws2.columns:
                max_length = 0
                column = col[0].column_letter  # Obtener la letra de la columna
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 4)
                ws2.column_dimensions[column].width = adjusted_width
                
                
            #########################################################################
            #________________________________Logo___________________________________
            #########################################################################
            
            
            # # Ruta de la imagen original
            # logo_path = r'.\Insumos\Logo_mundial_seguros.png'
            
            # # Cargar y redimensionar la imagen usando Pillow
            # logo = Image.open(logo_path)
            # new_width, new_height = 174, 58
            # logo_resized = logo.resize((new_width, new_height))
            
            # # # Guardar la imagen redimensionada temporalmente
            # # Generar un nombre único para la imagen redimensionada
            # unique_filename = f'Logo_mundial_seguros_{uuid.uuid4().hex}.png'
            # resized_logo_path = rf'.\Insumos\{unique_filename}'
            # logo_resized.save(resized_logo_path)
            
            ruta_logo = os.path.join(ruta_actual, 'data')
            
            
            # Agregar el logo a la hoja 'Resumen'
            logo_openpyxl_1 = OpenpyxlImage(os.path.join(ruta_logo, 'Logo_1.png')) 
            
            # Agregar el logo a la hoja 'Resumen'
            logo_openpyxl_2 = OpenpyxlImage(os.path.join(ruta_logo, 'Logo_2.png')) 
            
            # Agregar el logo a la hoja 'Resumen'
            ws1.add_image(logo_openpyxl_1, 'D2')  
            
            # Agregar el logo a la hoja 'Detalle'
            ws2.add_image(logo_openpyxl_2, 'B2')
            
            #________________________________________________________________________
            
            
            #########################################################################
            #______________________________Guardar Excel_____________________________
            #########################################################################
            
              
            # Crear la carpeta 'EXCEL' si no existe
            ruta_excel = os.path.join(ruta_actual, 'EXCEL')
            if not os.path.exists(ruta_excel):
                os.makedirs(ruta_excel)
            
            
            filename = f'Extracto_IPS_{nit}_2024_{mes}.xlsx'
            
            # Ruta completa del archivo Excel
            file_path = os.path.join(ruta_excel, filename)
            
            wb.save(file_path)
            
            #___________________Crear Archivo Destinatarios__________________________
            
            
            # Agregar la información al diccionario
            data_dict['ID RECLAMANTE'].append(nit)
            data_dict['RECLAMANTE'].append(entidad)
            data_dict['NOMBRE EXTRACTO IPS'].append(filename)
            data_dict['FECHA CORTE'].append(fecha_corte)
            data_dict['FECHA ULTIMO DIA MES'].append(ultimo_dia_mes)
            
            
            # Convertir el diccionario en un DataFrame
            df_destinatarios = pd.DataFrame(data_dict)
            
    
            #________________________________________________________________________
    
    
    
        #___________________Actualización Archivo Destinatarios______________________
         
           
        # Carpeta data
        ruta_data = os.path.join(ruta_actual, 'data')
            
    
        # Contacto Correos
        ruta_listado_correos = os.path.join(ruta_data, 'CONTACTO PRESTADORES.xlsx')
    
    
        # Lectura Contacto Correos
        df_correos = pd.read_excel(ruta_listado_correos)
    
    
        # Columnas selección df correos
        columnas_correos = ['Numero identificacion', 'CorreoPrincipal']
    
    
        # Selección de columnas
        df_correos_filtrado = df_correos[columnas_correos]
    
    
        df_correos_filtrado = df_correos_filtrado.drop_duplicates()
    
    
        # Conversión a numero
        df_correos_filtrado['Numero identificacion'] = pd.to_numeric(df_correos_filtrado['Numero identificacion'], errors='coerce')
    
    
        # Luego, convierte la columna a int64
        df_correos_filtrado['Numero identificacion'] = df_correos_filtrado['Numero identificacion'].astype('Int64')
    
    
        # print(df_correos_filtrado.info())
    
    
        # duplicados = df_correos_filtrado[df_correos_filtrado.duplicated(subset='Numero identificacion')]
    
    
        #___________________Cruce Info Destinatarios Vs Correos______________________
    
    
        df = pd.merge(df_destinatarios, df_correos_filtrado, 
                      how = 'left', 
                      left_on = 'ID RECLAMANTE',
                      right_on='Numero identificacion')
    
            
        # nombre archivo excel
        nombre_archivo_destinatarios = os.path.join(ruta_data, 'Destinatarios.xlsx')
    
            
        # Guardar el DataFrame en un archivo de Excel
        df.to_excel(nombre_archivo_destinatarios, index=False)
    
    
        print(f" >>> Extracto(s) generados correctamente!!!. \nRevisar resultado(s): {ruta_excel}.")
        
        progreso1.set(100)  # Completa la barra de progreso
        
        # fin = time.time()  # Tiempo de fin
        # tiempo = (fin - inicio) / 60
        
    except Exception as e:
        messagebox.showerror("Error", f"Error al generar los extractos: {str(e)}")
    


def enviar_correos():
    
    try:
        # inicio2 = time.time()
        # --------------- Configuración de la cuenta de Outlook ----------------- #
        outlook = win32com.client.Dispatch('outlook.application')
        namespace = outlook.GetNamespace("MAPI")
        cuenta_generica = None
    
        # Buscar la cuenta genérica
        for acc in namespace.Accounts:
            if acc.SmtpAddress == 'notificacionessoatmu@segurosmundial.com.co':
                cuenta_generica = acc
                break
    
        if cuenta_generica is None:
            raise Exception("No se encontró la cuenta genérica.")
         
        print(cuenta_generica) 
         
        # ----------------------------------------------------------------------- #
    
    
        def leer_destinatarios(filename):
            """
            Lee la lista de destinatarios desde un archivo Excel.
            
            Parámetros:
            filename (str): El nombre del archivo Excel.
    
            Retorna:
            DataFrame de pandas con la lista de destinatarios.
            """
            try:
                destinatarios = pd.read_excel(filename) 
                # print("Lista de destinatarios leída correctamente.")
                return destinatarios
            except Exception as e:
                print(f"Error al leer {filename}: {e}")
                raise
    
    
        def es_email_valido(email):
            """
            Valida si una dirección de correo electrónico es válida.
    
            Parámetros:
            email (str): La dirección de correo electrónico a validar.
    
            Retorna:
            bool: True si el correo es válido, False en caso contrario.
            """
            if pd.isna(email):
                return False
            regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
            return re.match(regex, email) is not None
    
    
        def generar_cuerpo_mensaje(entidad, fecha_ultimo_dia, fecha_corte):
            """
            Genera el cuerpo del mensaje de correo.
    
            Parámetros:
            entidad (str): El nombre de la entidad.
            fecha_reclamaciones (str): La fecha de las reclamaciones (mes y año).
            fecha_corte (str): La fecha de corte.
    
            Retorna:
            str: El cuerpo del mensaje de correo.
            """
            cuerpo_mensaje = f"""
            <p>Entidad: <strong><em>{entidad}</em></strong></p>
            
            <p>Seguros Mundial presenta en el archivo adjunto el extracto de información con el estado de las últimas reclamaciones por factura recibidas en {fecha_corte}.</p>
    
            <p>La información presentada está con corte del {fecha_ultimo_dia}. Únicamente se incluyen datos de reclamaciones con cargo a las pólizas de SOAT de la Aseguradora.</p>
    
            <p>Queremos recordarte que, a partir del 4 de junio de 2024, todas las solicitudes de cartera y conciliación SOAT, deben ser radicadas exclusivamente a través de nuestra página web <a href="https://sf.segurosmundial.com.co/fspq/s/">https://sf.segurosmundial.com.co/fspq/s/</a>. Es fundamental que tengas en cuenta que nuestro buzón <a href="mailto:carterasoat@segurosmundial.com.co">carterasoat@segurosmundial.com.co</a> no permitirá recepcionar más solicitudes.</p>
    
            <p>Para asegurar que tu solicitud sea procesada de manera eficiente y segura, te recomendamos utilizar nuestra plataforma. Puedes acceder al formulario en el siguiente enlace: <a href="https://sf.segurosmundial.com.co/fspq/s/">https://sf.segurosmundial.com.co/fspq/s/</a>.</p>
    
            <p>Saludos cordiales,<br>
            El equipo de Seguros Mundial</p>
            """
            return cuerpo_mensaje
    
    
        def enviar_correo(destinatario_email, nit, entidad, extracto_ips_excel, fecha_ultimo_dia, fecha_corte):
            """
            Envía un correo electrónico a un destinatario con archivos adjuntos específicos.
    
            Parámetros:
            destinatario_email (str): El correo electrónico del destinatario.
            nit (str): El NIT del destinatario.
            entidad (str): El nombre de la entidad.
            extracto_ips_excel (str): Nombre archivo excel con la información del extracto ips
            fecha_corte (str): La fecha de corte de la información.
    
            Retorna:
            dict: Diccionario con el resultado del envío del correo.
            """
           
            
            try:
            
                if not es_email_valido(destinatario_email):
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad, 'estado': 'Correo Inválido'}
            
          
                # Suponiendo que el nombre del archivo sigue el formato 'Extracto_IPS_<número>_YYYY_MM.xlsx'
                nombre_archivo = os.path.basename(extracto_ips_excel)  # Obtener solo el nombre del archivo
                partes = nombre_archivo.split('_')  # Dividir el nombre del archivo en partes
                fecha = f"{partes[-2]}_{partes[-1]}"  # Extraer la parte 'YYYY_MM' del nombre del archivo
                
                mail = outlook.CreateItem(0)
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, cuenta_generica))  # Usar la cuenta genérica
                mail.To = destinatario_email
                mail.Subject = f"Extracto IPS {entidad} {fecha}"
                cuerpo_mensaje = generar_cuerpo_mensaje(entidad, fecha_ultimo_dia, fecha_corte)
                mail.HTMLBody = cuerpo_mensaje
               
                
                # Definir las rutas de los archivos adjuntos
                base_path = os.getcwd()
                excel_filename = extracto_ips_excel
                pdf_folder_path = os.path.join(base_path, "PDF")
                excel_path = os.path.join(base_path, "EXCEL", excel_filename)
                
                if not os.path.exists(excel_path):
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad, 'estado': f"Excel no encontrado: {excel_path}"}
    
                # Adjuntar todos los archivos PDF en la carpeta
                adjuntos_pdf = 0
                if os.path.exists(pdf_folder_path):
                    for pdf_file in os.listdir(pdf_folder_path):
                        if pdf_file.endswith(".pdf"):
                            pdf_path = os.path.join(pdf_folder_path, pdf_file)
                            mail.Attachments.Add(pdf_path)
                            adjuntos_pdf += 1
                else:
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad,'estado': f"PDFs no encontrados en {pdf_folder_path}"}
                
                # Adjuntar archivo Excel
                adjuntos_excel = 0
                if os.path.exists(excel_path):
                    mail.Attachments.Add(excel_path)
                    adjuntos_excel += 1
                else:
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad, 'estado': f"Excel no encontrado: {excel_path}"}
    
                # Enviar el correo solo si hay archivos adjuntos
                if adjuntos_pdf > 0 and adjuntos_excel > 0:
                    mail.Send()
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad,'estado': 'Enviado'}  # Correo enviado
                else:
                    return {'email': destinatario_email, 'nit': nit, 'entidad': entidad, 'estado': 'No Enviado (sin adjuntos)'}  # No se envió el correo debido a falta de adjuntos
    
            except Exception as e:
                return {'email': destinatario_email, 'nit': nit, 'entidad': entidad,'estado': f"Error: {e}"}  # Error al enviar el correo
            
            
    
    
        # Base directorio actual
        base_path = os.getcwd()
    
        # Leer la lista de destinatarios
        destinatarios = leer_destinatarios(os.path.join(base_path, 'data', 'Destinatarios.xlsx'))
    
        # Filtrar destinatarios con correos nulos
        destinatarios_nulos = destinatarios[destinatarios['CorreoPrincipal'].isna()]
    
    
        # Filtrar destinatarios con correos válidos
        destinatarios = destinatarios[destinatarios['CorreoPrincipal'].notna()]
        # destinatarios.info()
    
    
        # Crear una lista para registrar los resultados del envío
        resultados = []
    
        # Iterar sobre la lista de destinatarios y enviar los correos
        for _, row in destinatarios.iterrows():
            resultado = enviar_correo(
                row['CorreoPrincipal'], 
                row['ID RECLAMANTE'], 
                row['RECLAMANTE'], 
                row['NOMBRE EXTRACTO IPS'], 
                row['FECHA ULTIMO DIA MES'], 
                row['FECHA CORTE'] 
            ) 
            resultados.append(resultado)
    
        # Guardar los resultados en un archivo CSV para su revisión
        resultados_df = pd.DataFrame(resultados)
        resultados_df.to_csv(os.path.join(base_path,'resultados_envio_correos.csv'), index=False)
        destinatarios_nulos.to_csv(os.path.join(base_path,'resultado_destinatarios_nulos.csv'), index=False)
    
    
    
        print("Proceso de envío de correos completado. Revisa el archivo resultados_envio_correos.csv para más detalles.")
        
        progreso2.set(100)  # Completa la barra de progreso
        
        # fin2 = time.time()  # Tiempo de fin
        # tiempo = (fin2 - inicio2) / 60
        
    except Exception as e:
        messagebox.showerror("Error", f"Error al enviar los correos: {str(e)}")



def actualizar_labels():
    # Leer el archivo CSV y actualizar las etiquetas
    df = pd.read_csv('resultados_envio_correos.csv')
    total_envios = len(df)
    fallidos = len(df[df['estado'] != 'Enviado'])
    total_envios = max(0, total_envios - fallidos)  # Asegurar que no sea negativo
    enviados_label.config(text=f"Total Correos Enviados: {total_envios}")
    fallidos_label.config(text=f"Total Correos Fallidos: {fallidos}")



def ejecutar_scripts():
    
    status_label.config(text="Estado: Ejecutando...")
    enviados_label.config(text="Total Correos Enviados: 0")
    fallidos_label.config(text="Total Correos Fallidos: 0")
    
    progreso1.set(0)
    progreso2.set(0)
    
        
    inicio1 = time.time()
    generar_extractos()
    tiempo1 = time.time() - inicio1
    status_label.config(text=f"Generación de Extractos Correctamente. \nTiempo: {tiempo1:.2f} minutos.")
    cantidad_extractos = contar_archivos_excel()
    
    # Mostrar mensaje script 1
    messagebox.showinfo("Generación de Extractos", f"Total: {cantidad_extractos} Extractos generados.")
    
    time.sleep(5)
    
    inicio2 = time.time()
    enviar_correos()
    tiempo2 = time.time() - inicio2
    status_label.config(text=f"Envío de Correos Exitosamente. Tiempo: {tiempo2:.2f} minutos.")
    cantidad_destinatarios = contar_destinatarios()
    
    # Mostrar mensaje script 2
    messagebox.showinfo("Envío de Correos", f"Total: {cantidad_destinatarios} Destinatarios.")
    
    #Esperar un tiempo fijo antes de actualizar las etiquetas
    root.after(5000, actualizar_labels())  # Espera 5 segundos antes de actualizar
    
    tiempo_total = tiempo1 + tiempo2
    
    
    # Mostrar mensaje de proceso finalizado
    messagebox.showinfo("Proceso Finalizado", f"Tiempo total de ejecución: {tiempo_total:.2f} minutos")
    


# Configuración de la ventana principal
root = tk.Tk()
root.title("Generación y Envío de Extractos IPS")
root.geometry("415x380")
root.resizable(False, False)

# ------------------------------ Background ------------------------------- #
canvas = tk.Canvas(root, width=415, height=380)
canvas.pack(fill="both", expand=True)

path_imagen = os.path.join(os.getcwd(), 'data', 'BG.jpg')
imagen_pil = Image.open(path_imagen)

def resize_image(event=None):
    width = root.winfo_width()
    height = root.winfo_height()
    nueva_imagen = imagen_pil.resize((width, height), Image.LANCZOS)
    photo = ImageTk.PhotoImage(nueva_imagen)
    canvas.create_image(0, 0, image=photo, anchor="nw")
    canvas.image = photo

root.after(100, resize_image)

frame = ttk.Frame(canvas, padding="10", style='TFrame')
canvas.create_window(200, 200, window=frame)

# ------------------------------------------------------------------------- #
frame = tk.Frame(canvas, bg='white', bd=2)
frame.place(relx=0.5, rely=0.5, anchor="center")

title_label = tk.Label(frame, text="Generación y Envío de Extractos IPS", font=('Helvetica', 16, 'bold'))
title_label.pack(pady=40)

run_scripts_button = tk.Button(frame, text="Iniciar Proceso Completo", font=('Helvetica', 12), bg='#4CAF50', fg='white', command=ejecutar_scripts)
run_scripts_button.pack(padx=10, pady=10)

# Barras de progreso
progreso1 = tk.DoubleVar()
progreso2 = tk.DoubleVar()
progress_bar1 = ttk.Progressbar(frame, variable=progreso1, maximum=100)
progress_bar1.pack(padx=10, pady=10)

progress_bar2 = ttk.Progressbar(frame, variable=progreso2, maximum=100)
progress_bar2.pack(padx=10, pady=10)

status_label = tk.Label(frame, text="Estado: Esperando", bg='white', font=('Helvetica', 13))
status_label.pack(padx=10, pady=10)

enviados_label = tk.Label(frame, text="Total Correos Enviados: 0", bg='white', font=('Helvetica', 12, 'bold'))
enviados_label.pack(padx=10, pady=10)

fallidos_label = tk.Label(frame, text="Total Correos Fallidos: 0", bg='white', font=('Helvetica', 12, 'bold'))
fallidos_label.pack(padx=10, pady=10)

root.mainloop()